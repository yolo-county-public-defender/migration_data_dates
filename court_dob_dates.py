import pandas as pd
import traceback
from openpyxl.styles import Alignment
from openpyxl.styles.numbers import FORMAT_DATE_XLSX14, NumberFormat
from datetime import datetime

def process_dates_and_times():
    try:
        print("Reading Excel file...")
        # Read all sheets
        excel_file = pd.read_excel('migrationdatamodified.xlsx', sheet_name=None)
        
        # Process People sheet - Column F (DOB)
        if 'People' in excel_file:
            df_people = excel_file['People'].copy()
            df_people.iloc[:, 5] = pd.to_datetime(df_people.iloc[:, 5], errors='coerce').dt.date
            excel_file['People'] = df_people
            
        # Process Court Date sheet - Columns C, D, E
        if 'Court Date' in excel_file:
            df_court = excel_file['Court Date'].copy()
            # Convert column C to date
            df_court.iloc[:, 2] = pd.to_datetime(df_court.iloc[:, 2], errors='coerce').dt.date
            # Convert columns D and E to time
            df_court.iloc[:, 3] = pd.to_datetime(df_court.iloc[:, 3], errors='coerce').dt.time
            df_court.iloc[:, 4] = pd.to_datetime(df_court.iloc[:, 4], errors='coerce').dt.time
            excel_file['Court Date'] = df_court
        
        print("Saving modified Excel file...")
        with pd.ExcelWriter('migrationdatamodified.xlsx', engine='openpyxl') as writer:
            for sheet_name, df in excel_file.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                if sheet_name == 'People':
                    worksheet = writer.sheets['People']
                    # Format column F (DOB) cells
                    for row in worksheet.iter_rows(min_row=2, min_col=6, max_col=6):
                        for cell in row:
                            cell.number_format = FORMAT_DATE_XLSX14
                            
                elif sheet_name == 'Court Date':
                    worksheet = writer.sheets['Court Date']
                    # Format column C (date) cells
                    for row in worksheet.iter_rows(min_row=2, min_col=3, max_col=3):
                        for cell in row:
                            cell.number_format = FORMAT_DATE_XLSX14
                    
                    # Format columns D and E (time) cells with 24-hour format
                    for row in worksheet.iter_rows(min_row=2, min_col=4, max_col=5):
                        for cell in row:
                            cell.number_format = 'HH:mm'
        
        print("Excel file has been processed successfully!")
        
    except Exception as e:
        print(f"\nAn error occurred: {str(e)}")
        print("Full error:")
        print(traceback.format_exc())

if __name__ == "__main__":
    process_dates_and_times()